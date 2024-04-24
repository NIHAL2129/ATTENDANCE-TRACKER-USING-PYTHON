import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

#LOADING THE EXCEL SHEET
book = openpyxl.load_workbook('attendance.xlsx')
sheet = book['Sheet1']

r = sheet.max_row
c = sheet.max_column
l1 = []
l2 = ""
l3 = []
staff_mails = ['xxxx@gmail.com']
m1 = "warning!!! you can take only one more day leave for PYTHON class"
m2 = "warning!!! you can take only one more day leave for JAVA class"
m3 = "warning!!! you can take only one more day leave for CPP class"

def savefile():
    book.save('attendance.xlsx')
    print("saved!")

def check(no_of_days, row_num, b):
    global l1, l2, l3
    for student in range(0, len(row_num)):
        if no_of_days[student] == 2:
            if b == 1:
                l1.append(sheet.cell(row=row_num[student], column=2).value)
                mailstu(l1, m1)
            elif b == 2:
                l1.append(sheet.cell(row=row_num[student], column=2).value)
                mailstu(l1, m2)
            else:
                l1.append(sheet.cell(row=row_num[student], column=2).value)
                mailstu(l1, m3)
        elif no_of_days[student] > 2:
            if b == 1:
                l2 += str(sheet.cell(row=row_num[student], column=1).value)
                l3.append(sheet.cell(row=row_num[student], column=2).value)
                subject = "PYTHON"
            elif b == 2:
                l2 += str(sheet.cell(row=row_num[student], column=1).value)
                l3.append(sheet.cell(row=row_num[student], column=2).value)
                subject = "JAVA"
            else:
                l2 += str(sheet.cell(row=row_num[student], column=1).value)
                l3.append(sheet.cell(row=row_num[student], column=2).value)
                subject = "CPP"
        if l2 != "" and len(l3) != 0:
            msg1 = "you have lack of attendance in " + subject + " !!!"
            msg2 = "the following students have lack of attendance in your subject : "+l2
            mailstu(l3, msg1)
            staff_id = staff_mails[b-1]
            mailstaff(staff_id, msg2)

def mailstu(li, msg):
    from_id = 'xxxx@gmail.com'
    pwd = 'enter the 16-charachter app password'
    s = smtplib.SMTP('smtp.gmail.com', 587, timeout=120)
    s.starttls()
    s.login(from_id, pwd)
    for i in range(0, len(li)):
        to_id = li[i]
        message = MIMEMultipart()
        message['Subject'] = 'Attendance report'
        message.attach(MIMEText(msg, 'plain'))
        content = message.as_string()
        s.sendmail(from_id, to_id, content)
    s.quit()
    print("mail sent to students successfully")

def mailstaff(mail_id, msg):
    from_id = 'xxxx@gmail.com'
    pwd = 'enter the 16-charachter app password'
    to_id = mail_id
    message = MIMEMultipart()
    message['Subject'] = 'Lack of attendance report'
    message.attach(MIMEText(msg, 'plain'))
    s = smtplib.SMTP('smtp.gmail.com', 587, timeout=120)
    s.starttls()
    s.login(from_id, pwd)
    content = message.as_string()
    s.sendmail(from_id, to_id, content)
    s.quit()
    print('Mail Sent to staff')

while True:
    print("1--->SUBJECT1\n2--->SUBJECT2\n3--->SUBJECT3")
    y = int(input("enter subject :"))
    no_of_absentees = int(input('no.of.absentees :'))
    if no_of_absentees > 1:
        x = list(map(int, (input('roll nos :').split(' '))))
    else:
        x = [int(input('roll no :'))]
    row_num = []
    no_of_days = []
    for student in x:
        for i in range(2, r+1):
            if y == 1:
                if sheet.cell(row=i, column=1).value == student:
                    m = sheet.cell(row=i, column=3).value
                    m += 1
                    sheet.cell(row=i, column=3).value = m
                    savefile()
                    no_of_days.append(m)
                    row_num.append(i)
            elif y == 2:
                if sheet.cell(row=i, column=1).value == student:
                    m = sheet.cell(row=i, column=4).value
                    m += 1
                    sheet.cell(row=i, column=4).value = m
                    no_of_days.append(m)
                    row_num.append(i)
            elif y == 3:
                if sheet.cell(row=i, column=1).value == student:
                    m = sheet.cell(row=i, column=5).value
                    m += 1
                    sheet.cell(row=i, column=5).value = m
                    row_num.append(i)
                    no_of_days.append(m)
    check(no_of_days, row_num, y)
    resp = int(input('another subject ? 1---->yes 0--->no'))
